import requests
import openpyxl
from datetime import datetime

# 🔁 Replace with your Codeforces handle
handle = "type your handle here"

# 📄 Excel file name
filename = "codeforces_log.xlsx"

# 🧾 Try to load existing file or make a new one
try:
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Date", "Problem Name", "Rating", "Tags", "Link", "Verdict", "Personal Note"])

# 🌐 Get submissions from Codeforces API
url = f"https://codeforces.com/api/user.status?handle={handle}&from=1&count=250"
res = requests.get(url).json()

if res["status"] != "OK":
    print("❌ Couldn’t fetch submissions. Check handle or internet.")
    exit()

problems_added = 0

for sub in res["result"]:
    if sub.get("verdict") != "OK":
        continue

    prob = sub["problem"]
    name = prob.get("name", "N/A")
    rating = prob.get("rating", "N/A")
    tags = ", ".join(prob.get("tags", []))
    link = f"https://codeforces.com/problemset/problem/{prob['contestId']}/{prob['index']}"
    timestamp = datetime.fromtimestamp(sub["creationTimeSeconds"]).strftime("%Y-%m-%d %H:%M")

    # ❌ Avoid duplicates
    already_logged = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == name and row[4] == link:
            already_logged = True
            break

    if not already_logged:
        sheet.append([timestamp, name, rating, tags, link, "OK", ""])  # Leave notes blank
        problems_added += 1

wb.save(filename)
print(f"✅ Logged {problems_added} new problem(s) to {filename}")
